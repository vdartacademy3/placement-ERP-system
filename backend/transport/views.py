from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from datetime import date, timedelta
import qrcode
import io
import os
import random
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    Route, Stop, Driver, Bus, StudentProfile, TransportPass,
    Trip, TripAttendance, Maintenance, FuelLog, Notification
)
from .serializers import (
    RouteSerializer, StopSerializer, DriverSerializer, BusSerializer,
    StudentProfileSerializer, TransportPassSerializer, TripSerializer,
    TripAttendanceSerializer, MaintenanceSerializer, FuelLogSerializer,
    NotificationSerializer, RegisterSerializer, UserSerializer
)


# ─── Auth ────────────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'User registered successfully'}, status=201)
    return Response(serializer.errors, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me(request):
    user = request.user
    role = 'manager' if user.is_staff else 'student'
    if hasattr(user, 'driver'):
        role = 'driver'
    data = UserSerializer(user).data
    data['role'] = role
    return Response(data)


# ─── Dashboard ───────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    today = date.today()
    total_buses = Bus.objects.count()
    active_buses = Bus.objects.filter(status='running').count()
    total_drivers = Driver.objects.filter(is_active=True).count()
    students_using_transport = StudentProfile.objects.filter(route__isnull=False).count()
    todays_trips = Trip.objects.filter(date=today).count()
    delayed_trips = Trip.objects.filter(date=today, status='delayed').count()
    maintenance_due = Bus.objects.filter(
        next_service_date__lte=today + timedelta(days=20)
    ).count()
    fuel_this_month = FuelLog.objects.filter(
        filled_at__month=today.month, filled_at__year=today.year
    ).aggregate(total=Sum('liters'))['total'] or 0
    live_buses = Bus.objects.filter(status='running').count()

    # Monthly trip data (last 6 months)
    monthly_data = []
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=30 * i)
        count = Trip.objects.filter(
            date__month=month_date.month, date__year=month_date.year
        ).count()
        monthly_data.append({
            'month': month_date.strftime('%b'),
            'trips': count,
            'simulated': random.randint(40, 120)
        })

    # Route popularity
    route_popularity = Route.objects.annotate(
        student_count=Count('studentprofile')
    ).values('name', 'student_count').order_by('-student_count')[:6]

    # Fuel expense last 6 months
    fuel_data = []
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=30 * i)
        cost = FuelLog.objects.filter(
            filled_at__month=month_date.month, filled_at__year=month_date.year
        ).aggregate(total=Sum('total_cost'))['total'] or random.randint(5000, 20000)
        fuel_data.append({'month': month_date.strftime('%b'), 'cost': float(cost)})

    return Response({
        'stats': {
            'total_buses': total_buses,
            'active_buses': active_buses,
            'total_drivers': total_drivers,
            'students_using_transport': students_using_transport,
            'todays_trips': todays_trips,
            'delayed_trips': delayed_trips,
            'maintenance_due': maintenance_due,
            'fuel_this_month': round(fuel_this_month, 2),
            'live_buses': live_buses,
        },
        'monthly_trips': monthly_data,
        'route_popularity': list(route_popularity),
        'fuel_expense': fuel_data,
    })


# ─── Route ViewSet ────────────────────────────────────────────────────────────

class RouteViewSet(viewsets.ModelViewSet):
    queryset = Route.objects.all()
    serializer_class = RouteSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name']

    @action(detail=True, methods=['post'])
    def reorder_stops(self, request, pk=None):
        route = self.get_object()
        stop_orders = request.data.get('stop_orders', [])
        for item in stop_orders:
            Stop.objects.filter(id=item['id'], route=route).update(order=item['order'])
        return Response({'message': 'Stops reordered'})


# ─── Stop ViewSet ─────────────────────────────────────────────────────────────

class StopViewSet(viewsets.ModelViewSet):
    queryset = Stop.objects.all()
    serializer_class = StopSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['route']
    search_fields = ['name', 'landmark']


# ─── Driver ViewSet ───────────────────────────────────────────────────────────

class DriverViewSet(viewsets.ModelViewSet):
    queryset = Driver.objects.all()
    serializer_class = DriverSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'license_number', 'mobile']

    @action(detail=True, methods=['get'])
    def trip_history(self, request, pk=None):
        driver = self.get_object()
        trips = Trip.objects.filter(driver=driver).order_by('-date')[:20]
        return Response(TripSerializer(trips, many=True).data)

    @action(detail=True, methods=['get'])
    def today_route(self, request, pk=None):
        driver = self.get_object()
        today_trip = Trip.objects.filter(driver=driver, date=date.today()).first()
        if today_trip:
            return Response(TripSerializer(today_trip).data)
        return Response({'message': 'No trip today'})


# ─── Bus ViewSet ──────────────────────────────────────────────────────────────

class BusViewSet(viewsets.ModelViewSet):
    queryset = Bus.objects.select_related('driver', 'route').all()
    serializer_class = BusSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'route', 'driver']
    search_fields = ['bus_number', 'registration_number']

    @action(detail=True, methods=['post'])
    def update_location(self, request, pk=None):
        bus = self.get_object()
        bus.latitude = request.data.get('latitude', bus.latitude)
        bus.longitude = request.data.get('longitude', bus.longitude)
        bus.save(update_fields=['latitude', 'longitude'])
        return Response({'latitude': bus.latitude, 'longitude': bus.longitude})

    @action(detail=False, methods=['get'])
    def live_locations(self, request):
        buses = Bus.objects.filter(status='running').values(
            'id', 'bus_number', 'latitude', 'longitude', 'status'
        )
        return Response(list(buses))


# ─── Student ViewSet ──────────────────────────────────────────────────────────

class StudentProfileViewSet(viewsets.ModelViewSet):
    queryset = StudentProfile.objects.select_related('user', 'route', 'stop').all()
    serializer_class = StudentProfileSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['route', 'stop', 'department', 'year']
    search_fields = ['user__first_name', 'user__last_name', 'roll_number', 'department']


# ─── Transport Pass ViewSet ───────────────────────────────────────────────────

class TransportPassViewSet(viewsets.ModelViewSet):
    queryset = TransportPass.objects.select_related('student', 'route', 'stop').all()
    serializer_class = TransportPassSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status', 'route']
    search_fields = ['student__roll_number', 'student__user__first_name']

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        transport_pass = self.get_object()
        transport_pass.status = 'approved'
        transport_pass.approved_at = timezone.now()
        transport_pass.valid_from = date.today()
        transport_pass.valid_until = date.today() + timedelta(days=180)
        # Generate QR code
        qr_data = f"PASS:{transport_pass.id}|STUDENT:{transport_pass.student.roll_number}|ROUTE:{transport_pass.route.name if transport_pass.route else 'N/A'}|VALID:{transport_pass.valid_until}"
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        filename = f"qr_{transport_pass.student.roll_number}.png"
        transport_pass.qr_code.save(filename, ContentFile(buffer.getvalue()), save=False)
        transport_pass.save()
        Notification.objects.create(
            title='Bus Pass Approved',
            message=f'Your transport pass has been approved. Valid until {transport_pass.valid_until}.',
            notification_type='general',
            recipient=transport_pass.student.user,
        )
        return Response(TransportPassSerializer(transport_pass, context={'request': request}).data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        transport_pass = self.get_object()
        transport_pass.status = 'rejected'
        transport_pass.save()
        Notification.objects.create(
            title='Bus Pass Rejected',
            message='Your transport pass application has been rejected. Please contact the transport office.',
            notification_type='general',
            recipient=transport_pass.student.user,
        )
        return Response({'message': 'Pass rejected'})


# ─── Trip ViewSet ─────────────────────────────────────────────────────────────

class TripViewSet(viewsets.ModelViewSet):
    queryset = Trip.objects.select_related('bus', 'driver', 'route').all()
    serializer_class = TripSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'bus', 'driver', 'route', 'date', 'trip_type']
    search_fields = ['bus__bus_number', 'driver__name']
    ordering_fields = ['date', 'scheduled_start']


# ─── Trip Attendance ViewSet ──────────────────────────────────────────────────

class TripAttendanceViewSet(viewsets.ModelViewSet):
    queryset = TripAttendance.objects.all()
    serializer_class = TripAttendanceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['trip', 'student', 'status']

    @action(detail=False, methods=['post'])
    def scan_qr(self, request):
        qr_data = request.data.get('qr_data', '')
        trip_id = request.data.get('trip_id')
        action_type = request.data.get('action', 'boarded')
        try:
            roll_number = qr_data.split('STUDENT:')[1].split('|')[0]
            student = StudentProfile.objects.get(roll_number=roll_number)
            trip = Trip.objects.get(id=trip_id)
            attendance, _ = TripAttendance.objects.get_or_create(trip=trip, student=student)
            if action_type == 'boarded':
                attendance.status = 'boarded'
                attendance.boarded_at = timezone.now()
            else:
                attendance.status = 'dropped'
                attendance.dropped_at = timezone.now()
            attendance.save()
            return Response({'message': f'Student {action_type} successfully', 'student': student.user.get_full_name()})
        except Exception as e:
            return Response({'error': str(e)}, status=400)


# ─── Maintenance ViewSet ──────────────────────────────────────────────────────

class MaintenanceViewSet(viewsets.ModelViewSet):
    queryset = Maintenance.objects.select_related('bus').all()
    serializer_class = MaintenanceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['bus', 'status', 'maintenance_type']
    search_fields = ['bus__bus_number', 'description']


# ─── Fuel Log ViewSet ─────────────────────────────────────────────────────────

class FuelLogViewSet(viewsets.ModelViewSet):
    queryset = FuelLog.objects.select_related('bus').all()
    serializer_class = FuelLogSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['bus']
    ordering_fields = ['filled_at']


# ─── Notification ViewSet ─────────────────────────────────────────────────────

class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        return Notification.objects.filter(
            Q(recipient=user) | Q(is_broadcast=True)
        )

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notif = self.get_object()
        notif.is_read = True
        notif.save()
        return Response({'message': 'Marked as read'})

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        self.get_queryset().update(is_read=True)
        return Response({'message': 'All marked as read'})

    @action(detail=False, methods=['post'])
    def broadcast(self, request):
        if not request.user.is_staff:
            return Response({'error': 'Permission denied'}, status=403)
        Notification.objects.create(
            title=request.data.get('title'),
            message=request.data.get('message'),
            notification_type=request.data.get('type', 'general'),
            is_broadcast=True,
        )
        return Response({'message': 'Broadcast sent'})


# ─── AI Features ─────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ai_delay_prediction(request):
    bus_id = request.query_params.get('bus_id')
    buses = Bus.objects.filter(status='running') if not bus_id else Bus.objects.filter(id=bus_id)
    predictions = []
    weather_conditions = ['Clear', 'Cloudy', 'Light Rain', 'Heavy Rain', 'Fog']
    traffic_levels = ['Low', 'Moderate', 'High', 'Very High']
    for bus in buses:
        weather = random.choice(weather_conditions)
        traffic = random.choice(traffic_levels)
        delay_score = 0
        if weather in ['Heavy Rain', 'Fog']:
            delay_score += 40
        elif weather == 'Light Rain':
            delay_score += 20
        if traffic == 'Very High':
            delay_score += 40
        elif traffic == 'High':
            delay_score += 25
        elif traffic == 'Moderate':
            delay_score += 10
        delay_score += random.randint(0, 20)
        predictions.append({
            'bus_id': bus.id,
            'bus_number': bus.bus_number,
            'prediction': 'Late' if delay_score > 50 else 'On Time',
            'confidence': min(95, 60 + random.randint(0, 35)),
            'weather': weather,
            'traffic': traffic,
            'estimated_delay_min': random.randint(5, 25) if delay_score > 50 else 0,
        })
    return Response(predictions)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ai_seat_availability(request):
    buses = Bus.objects.filter(status__in=['running', 'available'])
    data = []
    for bus in buses:
        today_trip = Trip.objects.filter(bus=bus, date=date.today()).first()
        boarded = 0
        if today_trip:
            boarded = today_trip.attendances.filter(status='boarded').count()
        else:
            boarded = random.randint(0, bus.capacity)
        remaining = max(0, bus.capacity - boarded)
        data.append({
            'bus_id': bus.id,
            'bus_number': bus.bus_number,
            'capacity': bus.capacity,
            'boarded': boarded,
            'remaining': remaining,
            'occupancy_percent': round((boarded / bus.capacity) * 100, 1) if bus.capacity else 0,
        })
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ai_maintenance_prediction(request):
    buses = Bus.objects.all()
    predictions = []
    for bus in buses:
        days_since_service = 0
        if bus.last_service_date:
            days_since_service = (date.today() - bus.last_service_date).days
        risk_score = min(100, (days_since_service / 90) * 50 + (bus.total_mileage_km / 10000) * 30 + random.randint(0, 20))
        days_to_maintenance = max(0, 20 - int(risk_score / 5))
        predictions.append({
            'bus_id': bus.id,
            'bus_number': bus.bus_number,
            'risk_level': 'High' if risk_score > 70 else 'Medium' if risk_score > 40 else 'Low',
            'risk_score': round(risk_score, 1),
            'days_to_maintenance': days_to_maintenance,
            'mileage': bus.total_mileage_km,
            'days_since_service': days_since_service,
            'recommendation': f'Bus may require maintenance within {days_to_maintenance} days' if risk_score > 40 else 'Bus is in good condition',
        })
    return Response(predictions)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ai_route_suggestion(request):
    routes = Route.objects.annotate(student_count=Count('studentprofile')).all()
    suggestions = []
    for route in routes:
        traffic_factor = random.choice(['Low', 'Moderate', 'High'])
        suggestions.append({
            'route_id': route.id,
            'route_name': route.name,
            'current_distance_km': route.distance_km,
            'suggested_distance_km': round(route.distance_km * random.uniform(0.85, 0.95), 1),
            'traffic': traffic_factor,
            'recommendation': 'Use alternate highway route' if traffic_factor == 'High' else 'Current route is optimal',
            'time_saving_min': random.randint(5, 15) if traffic_factor == 'High' else 0,
            'student_count': route.student_count,
        })
    return Response(suggestions)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ai_demand_analysis(request):
    routes = Route.objects.annotate(
        student_count=Count('studentprofile'),
        bus_count=Count('buses'),
    ).values('id', 'name', 'student_count', 'bus_count')
    data = []
    for r in routes:
        demand = 'High' if r['student_count'] > 30 else 'Medium' if r['student_count'] > 15 else 'Low'
        data.append({
            **r,
            'demand_level': demand,
            'crowding_index': round(r['student_count'] / max(r['bus_count'] * 40, 1) * 100, 1),
        })
    return Response(sorted(data, key=lambda x: x['student_count'], reverse=True))


# ─── Reports ──────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_pdf_report(request):
    report_type = request.query_params.get('type', 'daily')
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_transport_report.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph(f'Transport {report_type.title()} Report', styles['Title']))
    elements.append(Paragraph(f'Generated: {timezone.now().strftime("%d %b %Y %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 20))

    if report_type in ['daily', 'monthly']:
        trips = Trip.objects.all()[:50]
        data = [['Bus', 'Driver', 'Route', 'Type', 'Status', 'Date']]
        for t in trips:
            data.append([
                t.bus.bus_number if t.bus else '-',
                t.driver.name if t.driver else '-',
                t.route.name if t.route else '-',
                t.trip_type,
                t.status,
                str(t.date),
            ])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

    elif report_type == 'fuel':
        logs = FuelLog.objects.all()[:50]
        data = [['Bus', 'Liters', 'Cost/L', 'Total Cost', 'Date']]
        for f in logs:
            data.append([f.bus.bus_number, f.liters, str(f.cost_per_liter), str(f.total_cost), str(f.filled_at.date())])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

    elif report_type == 'maintenance':
        records = Maintenance.objects.all()[:50]
        data = [['Bus', 'Type', 'Scheduled', 'Status', 'Cost']]
        for m in records:
            data.append([m.bus.bus_number, m.maintenance_type, str(m.scheduled_date), m.status, str(m.cost)])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

    doc.build(elements)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_excel_report(request):
    report_type = request.query_params.get('type', 'daily')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{report_type.title()} Report'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')

    if report_type in ['daily', 'monthly']:
        headers = ['Bus Number', 'Driver', 'Route', 'Trip Type', 'Status', 'Date', 'Fuel Used (L)', 'Distance (km)']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for t in Trip.objects.all()[:100]:
            ws.append([
                t.bus.bus_number if t.bus else '-',
                t.driver.name if t.driver else '-',
                t.route.name if t.route else '-',
                t.trip_type, t.status, str(t.date),
                t.fuel_used_liters, t.distance_covered_km,
            ])

    elif report_type == 'fuel':
        headers = ['Bus Number', 'Liters', 'Cost/Liter', 'Total Cost', 'Odometer', 'Date']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        for f in FuelLog.objects.all()[:100]:
            ws.append([f.bus.bus_number, f.liters, float(f.cost_per_liter), float(f.total_cost), f.odometer_reading, str(f.filled_at.date())])

    elif report_type == 'maintenance':
        headers = ['Bus Number', 'Type', 'Description', 'Scheduled Date', 'Completed Date', 'Status', 'Cost']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
        for m in Maintenance.objects.all()[:100]:
            ws.append([m.bus.bus_number, m.maintenance_type, m.description, str(m.scheduled_date), str(m.completed_date or ''), m.status, float(m.cost)])

    elif report_type == 'student':
        headers = ['Roll Number', 'Name', 'Department', 'Year', 'Route', 'Stop', 'Pass Status']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        for s in StudentProfile.objects.select_related('user', 'route', 'stop').all()[:100]:
            pass_status = s.transport_pass.status if hasattr(s, 'transport_pass') else 'N/A'
            ws.append([s.roll_number, s.user.get_full_name(), s.department, s.year,
                       s.route.name if s.route else '-', s.stop.name if s.stop else '-', pass_status])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_transport_report.xlsx"'
    wb.save(response)
    return response
